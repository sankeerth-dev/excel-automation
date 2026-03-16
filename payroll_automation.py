import os
import shutil
import pandas as pd
from datetime import datetime
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from num2words import num2words
import calendar

def generate_dummy_data(filename):
    """Generates a sample attendance file if one doesn't exist to test the workflow."""
    data = {
        "Employee ID": [1, 2, 3],
        "Name of the Resource": ["Akash Totiger", "John Doe", "Jane Smith"],
        "Designation": ["SDE 1", "HR Manager", "QA Engineer"],
        "PAN": ["ABCD1234E", "WXYZ9876Q", "MNBQ4567P"],
        "Bank Name": ["HDFC Bank", "ICICI Bank", "SBI"],
        "Bank A/C No.": ["012345678901", "987654321098", "456123789012"],
        "P.F. A/C Number": ["MH/BAN/12345/000/111", "MH/BAN/98765/000/222", "MH/BAN/45612/000/333"],
        "UAN Number": ["100987654321", "100123456789", "100456789123"],
        "Payment": [37500, 50000, 45000],  # Monthly Master CTC Before Deductions
        "Basic": [15000, 20000, 18000],
        "House Rent Allowance": [6000, 8000, 7200],
        "Special Allowance": [16500, 22000, 19800],
        "Statutory Bonus": [0, 0, 0],
        "LTA Allowance": [0, 0, 0],
        "Other Earning 1": [0, 0, 0],
        "Other Earning 2": [0, 0, 0],
        "Other Earning 3": [0, 0, 0],
        "Total Earnings": [37500, 50000, 45000], 
        "Tax": [200, 500, 300],  # Tax / Deductions
        "Per day": [0, 0, 0], # To be calculated
        "Number of days": [31, 31, 31],
        "Number of days Attended": [22, 31, 20],
        "Total": [0, 0, 0], # To be calculated
        "Leave Taken": [0, 0, 0], # To be calculated
        "Comment": ["", "", "Sick Leave"]
    }
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)
    print(f"Generated sample template: {filename}")

def process_payroll(input_file, current_month, current_year):
    """Reads the attendance data, calculates the fields and saves the updated payroll sheet."""
    # Load Data
    df = pd.read_excel(input_file)
    
    # Strip newlines and collapse multiple spaces from columns for consistent matching
    df.columns = [re.sub(r'\s+', ' ', str(c).strip()) for c in df.columns]
    
    # Filter out completely empty rows or rows without a name
    df = df[df["Name of the Resource"].notnull()].copy()
    
    # Ensure missing expected columns exist to prevent crashes on older templates
    expected_cols = [
        "Employee ID", "Name of the Resource", "Designation",
        "PAN", "Bank Name", "Bank A/C No.", "P.F. A/C Number", "UAN Number",
        "Payment", "Basic", "House Rent Allowance", "Special Allowance",
        "Statutory Bonus", "LTA Allowance", "Other Earning 1",
        "Other Earning 2", "Other Earning 3", "Total Earnings",
        "Tax", "Number of days", "Number of days Attended", "Comment"
    ]
    numeric_defaults = ["Payment", "Basic", "House Rent Allowance", "Special Allowance", "Statutory Bonus", "LTA Allowance", "Other Earning 1", "Other Earning 2", "Other Earning 3", "Total Earnings", "Tax", "Number of days", "Number of days Attended"]
    
    for col in expected_cols:
        if col not in df.columns:
            df[col] = 0 if col in numeric_defaults else ""

    # Clean data (ensure numeric columns)
    for col in numeric_defaults:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    # Avoid Division by Zero
    df["Number of days"] = df["Number of days"].replace(0, 30) # Default to 30 if 0

    # Calculate fields
    df["Leave Taken"] = (df["Number of days"] - df["Number of days Attended"]).round(0)
    
    # Assuming 'Payment' is Annual CTC, calculate monthly base
    df["Monthly Base"] = df["Payment"] / 12
    df["Per day"] = df["Monthly Base"] / df["Number of days"]
    
    # Apply LOP Deduction Rule: if leaves > 3, deduct all; if leaves == 3, deduct 1; else 0
    def calculate_lop_days(leave):
        if leave > 3:
            return leave
        elif leave == 3:
            return 1
        else:
            return 0
    
    df["LOP Days Calculated"] = df["Leave Taken"].apply(calculate_lop_days)
    df["LOP Deduction"] = df["LOP Days Calculated"] * df["Per day"]
    
    # Total Earnings = Monthly Base - LOP Deduction
    # This represents the actual earning for the month before tax
    df["Total Earnings"] = df["Monthly Base"] - df["LOP Deduction"]
    
    # Professional Tax is now fixed at 200, Income Tax at 0
    df["Tax"] = 200.0
    df["Total"] = df["Total Earnings"] - df["Tax"]

    # Round up the calculated currencies to 2 decimal places
    df["Per day"] = df["Per day"].round(2)
    df["Total Earnings"] = df["Total Earnings"].round(2)
    df["Total"] = df["Total"].round(2)

    # The user asked not to change the columns in the payroll month sheet.
    # We will compute the fields but save the values directly to the copied master sheet.

    # Create output directory for the current month
    output_dir = f"{current_month}_{current_year}"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Save to a new file representing this month's processed payroll
    output_filename = os.path.join(output_dir, f"Payroll_{current_month}_{current_year}.xlsx")

    # Copy the master file to preserve its exact layout, colors, width, and validation
    shutil.copy2(input_file, output_filename)

    # Open the copied spreadsheet using openpyxl to update the cell values
    wb = load_workbook(output_filename)
    ws = wb.active

    # Read exactly what the columns are from the first row of the exact copied file
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col_idx).value
        # Sometimes headers have newlines, so we clean them up to match our pandas columns
        if header_val:
            col_map[header_val.strip().replace('\n', '')] = col_idx
            
    # Also record exact string headers from pandas dataframe to do the mapping robustly
    pd_col_to_excel_col = {}
    for pd_col in df.columns:
        # Search the excel headers
        for raw_ex_col, ex_col_idx in col_map.items():
            if raw_ex_col == pd_col.strip().replace('\n', ''):
                pd_col_to_excel_col[pd_col] = ex_col_idx
                break

    # We assume the data starts from row 2
    for r_idx, row in df.iterrows():
        excel_row = r_idx + 2
        for col_name, val in row.items():
            if col_name in pd_col_to_excel_col:
                col_idx = pd_col_to_excel_col[col_name]
                # openpyxl uses 1-based indexing
                ws.cell(row=excel_row, column=col_idx, value=val)

    wb.save(output_filename)
    print(f"✅ Processed Payroll Summary saved successfully to {output_filename}")
    
    # We return the dataframe to generate accurate payslips
    return df

def get_income_tax_percentage(annual_income):
    """Returns the tax percentage base on the provided annual income slabs."""
    if annual_income <= 400000:
        return 0, 0
    elif annual_income <= 800000:
        return 5, annual_income * 0.05
    elif annual_income <= 1200000:
        return 10, annual_income * 0.10
    elif annual_income <= 1600000:
        return 15, annual_income * 0.15
    elif annual_income <= 2000000:
        return 20, annual_income * 0.20
    elif annual_income <= 2400000:
        return 25, annual_income * 0.25
    else:
        return 30, annual_income * 0.30

def generate_salary_slips(df, current_month, current_year):
    """Generates the formatted salary slips for each employee in a separate Excel sheet and a PDF."""
    
    output_dir = f"{current_month}_{current_year}"
    pdf_dir = current_month
    if not os.path.exists(pdf_dir):
        os.makedirs(pdf_dir)

    # Initialize Excel application for PDF conversion
    try:
        import win32com.client
        excel_app = win32com.client.DispatchEx("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False
    except ImportError:
        excel_app = None
        print("\nNote: pywin32 is not installed. PDF salary slips will not be generated.")
        print("Run 'pip install pywin32' to enable PDF generation.\n")
    except Exception as e:
        excel_app = None
        print(f"\nNote: Could not initialize Excel application for PDF conversion: {e}\n")

    # Define Formatting Styles matching the image
    font_bold = Font(bold=True)
    font_company = Font(size=16, bold=True, color="000000")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    def get_border(l=False, r=False, t=False, b=False):
        bd = {}
        if l: bd['left'] = Side(style='thin')
        if r: bd['right'] = Side(style='thin')
        if t: bd['top'] = Side(style='thin')
        if b: bd['bottom'] = Side(style='thin')
        return Border(**bd)

    thin_border = get_border(True, True, True, True)
    side_border = get_border(True, True, False, False)

    # Mapping to calculate basic vs allowances for visualization
    for index, row in df.iterrows():
    
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
            
        import re
        raw_name = str(row["Name of the Resource"])
        sheet_name = re.sub(r'[\\/\*\?\[\]\:\n\r]', ' ', raw_name).strip()[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        payment = float(row["Payment"])
        days = float(row["Number of days"])
        attended = float(row["Number of days Attended"]) if pd.notnull(row["Number of days Attended"]) else 0
        tax = float(row["Tax"]) if pd.notnull(row["Tax"]) else 0
        leave = float(row["Leave Taken"]) if pd.notnull(row["Leave Taken"]) else 0
        # No longer using Earned Salary in the new formula logic
        
        # Read Earnings components from Attendance Master
        master_basic = float(row.get("Basic", 0.0))
        master_hra = float(row.get("House Rent Allowance", 0.0))
        master_special = float(row.get("Special Allowance", 0.0))
        master_bonus = float(row.get("Statutory Bonus", 0.0))
        master_lta = float(row.get("LTA Allowance", 0.0))
        master_other1 = float(row.get("Other Earning 1", 0.0))
        master_other2 = float(row.get("Other Earning 2", 0.0))
        master_other3 = float(row.get("Other Earning 3", 0.0))
        total_earnings_manual = float(row.get("Total Earnings", 0.0))
        
        # Fallback to the 40/16/44 rule based on Payment (Annual) if no components provided
        if master_basic == 0 and master_hra == 0 and master_special == 0 and payment > 0:
            monthly_payment = payment / 12
            master_basic = monthly_payment * 0.40
            master_hra = monthly_payment * 0.16
            master_special = monthly_payment - (master_basic + master_hra)
        
        # Gross Earnings (Sum of components for the A27 row)
        gross_earnings_sum = master_basic + master_hra + master_special + master_bonus + master_lta + master_other1 + master_other2 + master_other3
        
        # Total Earnings (For the Net Pay section)
        # We use the already calculated 'Total Earnings' from the dataframe 
        # which includes the LOP deduction logic (Monthly Base - LOP)
        total_earnings_final = float(row.get("Total Earnings", total_earnings_manual))
        
        # Calculate Income Tax based on final Total Earnings
        annual_income_for_tax = total_earnings_final * 12
        tax_pct, annual_tax = get_income_tax_percentage(annual_income_for_tax)
        monthly_income_tax = annual_tax / 12
        
        # Re-calculate LOP days for the footer display
        if leave > 3:
            lop_days_to_deduct = float(leave)
        elif leave == 3:
            lop_days_to_deduct = 1.0
        else:
            lop_days_to_deduct = 0.0
        
        total_deduction = 200.0 + monthly_income_tax # Professional Tax 200 + Dynamic Income Tax
        net_pay = total_earnings_final - total_deduction
        
        # Configure columns width
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 22
        
        # Write Headers (Company Name, Address)
        ws.merge_cells("A1:E1")
        ws["A1"] = "PATHAXIOM Pvt LTD"
        ws["A1"].font = font_company
        ws["A1"].alignment = align_left
        ws.row_dimensions[1].height = 55

        address = "PathAxiom Pvt. Ltd.\nNo. 332, Siddhaiah Puranik Road\n3rd Stage, 4th Block,\nShakthi Ganapathi Nagar,\nBasaveshwar Nagar,\nBengaluru, Karnataka - 560079\nIndia"
        ws.merge_cells("A2:C8")
        ws["A2"] = address
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

        # Top border box for address and logo
        for r in range(1, 9):
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = get_border(l=(c==1), r=(c==6), t=(r==1), b=(r==8))

        if os.path.exists("logo.png"):
            from openpyxl.drawing.image import Image as ExcelImage
            img = ExcelImage("logo.png")
            
            # Fit perfectly exactly into F1 cell boundaries
            # Column F width (22) is exactly ~160 pixels. Row 1 height (55) is exactly ~73 pixels
            img.width = 160
            img.height = 73
            
            img.anchor = 'F1'
            ws.add_image(img)

        # Payslip Month Header
        ws.merge_cells("A9:F9")
        ws["A9"] = f"Payslip for the month of {current_month} {current_year}"
        ws["A9"].font = font_bold
        ws["A9"].alignment = align_center
        ws.row_dimensions[9].height = 35
        for col in range(1, 7):
            ws.cell(row=9, column=col).border = get_border(l=(col==1), r=(col==6), b=True)

        # Employee Pay Summary Section
        ws.merge_cells("A10:C10")
        ws["A10"] = "Employee Pay Summary"
        ws["A10"].font = font_bold
        ws.merge_cells("D10:F10")
        for c in range(1, 4): ws.cell(row=10, column=c).border = thin_border
        for c in range(4, 7): ws.cell(row=10, column=c).border = thin_border
        
        summary_rows = [
            ("Employee Name", row["Name of the Resource"], "PAN", row.get("PAN", "")),
            ("Designation", row["Designation"], "Bank Name", row.get("Bank Name", "")),
            ("Employee ID", row.get("Employee ID", ""), "Bank A/C No.", row.get("Bank A/C No.", "")),
            ("Date of Joining", "01-12-2025", "P.F. A/C Number", row.get("P.F. A/C Number", "")),
            ("Department", "Development", "UAN Number", row.get("UAN Number", "")),
            ("Location", "Bangalore", "Days Worked", int(attended) if pd.notnull(attended) else 0),
            ("", "", "Pay Date (dd-mm-yyyy)", datetime.today().strftime("%d-%m-%Y"))
        ]
        
        row_num = 11
        for s in summary_rows:
            ws.cell(row=row_num, column=1, value=s[0]).border = thin_border
            
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3) # Merge B & C
            ws.cell(row=row_num, column=2, value=s[1]).border = thin_border
            ws.cell(row=row_num, column=2).alignment = align_left
            ws.cell(row=row_num, column=3).border = thin_border
            
            ws.cell(row=row_num, column=4, value=s[2]).border = thin_border
            
            ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6) # Merge E & F
            ws.cell(row=row_num, column=5, value=s[3]).border = thin_border
            ws.cell(row=row_num, column=5).alignment = align_center
            ws.cell(row=row_num, column=6).border = thin_border
            
            row_num += 1

        # Table Column Headers
        ws.cell(row=row_num, column=1, value="EARNINGS").font = font_bold
        ws.cell(row=row_num, column=2, value="Master").font = font_bold
        ws.cell(row=row_num, column=3, value="Earnings").font = font_bold
        ws.cell(row=row_num, column=4, value="Particulars").font = font_bold
        
        ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=5, value="Deductions").font = font_bold
        ws.cell(row=row_num, column=5).alignment = align_center
        
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = thin_border
            if col in [1, 4]:
                ws.cell(row=row_num, column=col).alignment = align_left

        row_num += 1

        # Make sure values precisely match the screenshot!
        rows_data = [
            ("Basic", 0.0, master_basic, f"Income Tax Deduction ({tax_pct}%)", monthly_income_tax), # Swap master and earnings back!
            ("House Rent Allowance", 0.0, master_hra, "Profession Tax", 200.0),
            ("Special Allowance", 0.0, master_special, "P.F.", 0.0),
            ("Statutory Bonus", 0.0, master_bonus, "Other Deduction 1", 0.0),
            ("LTA Allowance", 0.0, master_lta, "Other Deduction 2", 0.0),
            ("Other Earning 1", 0.0, master_other1, "Other Deduction 3", 0.0),
            ("Other Earning 2", 0.0, master_other2, "", ""),
            ("Other Earning 3", 0.0, master_other3, "", "")
        ]

        def format_currency(val):
            return f"₹{val:,.2f}" if isinstance(val, (int, float)) else val

        for ed in rows_data:
            for c in range(1, 7):
                ws.cell(row=row_num, column=c).border = side_border
                
            ws.cell(row=row_num, column=1, value=ed[0])
            ws.cell(row=row_num, column=2, value=format_currency(ed[1])).alignment = align_right
            ws.cell(row=row_num, column=3, value=format_currency(ed[2])).alignment = align_right
            
            if ed[3] != "":
                ws.cell(row=row_num, column=4, value=ed[3]).border = get_border(l=True, r=True)
            else:
                ws.cell(row=row_num, column=4, value=ed[3])
                
            ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)
            ws.cell(row=row_num, column=5, value=format_currency(ed[4]) if ed[4] != "" else "").alignment = align_right
            ws.cell(row=row_num, column=6).border = get_border(r=True)
            row_num += 1

        # Gross & Total Deductions
        ws.cell(row=row_num, column=1, value="Gross Earnings").font = font_bold
        ws.cell(row=row_num, column=2, value="")
        ws.cell(row=row_num, column=3, value=format_currency(gross_earnings_sum)).font = font_bold
        ws.cell(row=row_num, column=3).alignment = align_right
        
        ws.cell(row=row_num, column=4, value="Total Deductions").font = font_bold
        ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=5, value=format_currency(total_deduction)).font = font_bold
        ws.cell(row=row_num, column=5).alignment = align_right
        
        for c in range(1, 7):
            ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1

        # REIMBURSEMENTS Setup
        ws.cell(row=row_num, column=1, value="REIMBURSEMENTS").font = font_bold
        for c in range(1, 7): ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1

        for r_name in ["Reimbursement 1", "Reimbursement 2"]:
            for c in range(1, 7): ws.cell(row=row_num, column=c).border = side_border
            ws.cell(row=row_num, column=1, value=r_name)
            ws.cell(row=row_num, column=2, value="₹0.00").alignment = align_right
            ws.cell(row=row_num, column=3, value="₹0.00").alignment = align_right
            row_num += 1

        ws.cell(row=row_num, column=1, value="Total Reimbursements").font = font_bold
        ws.cell(row=row_num, column=2, value="")
        ws.cell(row=row_num, column=3, value="₹0.00").font = font_bold
        ws.cell(row=row_num, column=3).alignment = align_right
        for c in range(1, 7): ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1

        # NETPAY SECTION
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        ws.cell(row=row_num, column=1, value="NETPAY").font = font_bold
        
        ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=4, value="AMOUNT (Total Earnings)").font = font_bold
        ws.cell(row=row_num, column=4).alignment = align_left
        for c in range(1, 7): ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1

        netpays = [
             ("Total Earnings", format_currency(total_earnings_final)),
             ("Total Deductions", format_currency(total_deduction)),
             ("Total Reimbursements", "₹0.00")
        ]
        
        for n in netpays:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
            ws.cell(row=row_num, column=1, value=n[0])
            ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num, end_column=6)
            ws.cell(row=row_num, column=4, value=n[1]).alignment = align_right
            for c in range(1, 7): ws.cell(row=row_num, column=c).border = thin_border
            row_num += 1

        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
        ws.cell(row=row_num, column=1, value="Total Net Payable").alignment = align_right
        ws.cell(row=row_num, column=6, value=format_currency(net_pay)).font = font_bold
        ws.cell(row=row_num, column=6).alignment = align_right
        for c in range(1, 7): ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1

        # Amount in Words
        safe_net_pay = int(net_pay) if pd.notnull(net_pay) else 0
        words = num2words(safe_net_pay, lang='en_IN').capitalize() + " rupees only"
        ws.cell(row=row_num, column=1, value="").border = thin_border
        
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
        ws.cell(row=row_num, column=2, value="Amount in Words").font = font_bold
        ws.cell(row=row_num, column=2).alignment = align_center
        ws.cell(row=row_num, column=2).border = thin_border
        ws.cell(row=row_num, column=3).border = thin_border
        
        ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=4, value=f"( {words} )").font = font_bold
        ws.cell(row=row_num, column=4).alignment = align_center
        ws.cell(row=row_num, column=4).border = get_border(t=True, l=True, b=True)
        ws.cell(row=row_num, column=5).border = get_border(t=True, b=True)
        ws.cell(row=row_num, column=6).border = get_border(t=True, r=True, b=True)
        row_num += 1

        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=1, value="**Total Net Payable = Gross Earnings - Total Deductions + Total Reimbursements").alignment = align_center
        
        # Add outer borders for footer
        for c in range(1, 7):
            ws.cell(row=row_num, column=c).border = get_border(l=(c==1), r=(c==6))
        row_num += 1
        
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        # Add NaN safety for display
        display_leave = int(leave) if (pd.notnull(leave) and leave > 0) else 0
        display_lop = int(lop_days_to_deduct) if (pd.notnull(lop_days_to_deduct) and lop_days_to_deduct > 0) else 0
        ws.cell(row=row_num, column=1, value=f"L.O.P. Days : {display_leave if display_leave > 0 else ''} (Deducted Days: {display_lop if display_lop > 0 else 'None'})").alignment = align_center
        for c in range(1, 7):
            ws.cell(row=row_num, column=c).border = get_border(l=(c==1), r=(c==6), b=True)

        employee_filename = os.path.join(output_dir, f"Salary_Slip_{sheet_name}_{current_month}_{current_year}.xlsx")
        wb.save(employee_filename)
        
        # Add PDF Conversion
        if excel_app:
            try:
                abs_excel_path = os.path.abspath(employee_filename)
                pdf_filename = os.path.join(pdf_dir, f"Salary_Slip_{sheet_name}_{current_month}_{current_year}.pdf")
                abs_pdf_path = os.path.abspath(pdf_filename)
                
                # Open workbook in Excel
                excel_wb = excel_app.Workbooks.Open(abs_excel_path)
                ws_excel = excel_wb.ActiveSheet
                
                # Page setup to ensure it fits on one page exactly like the Excel layout
                ws_excel.PageSetup.Zoom = False
                ws_excel.PageSetup.FitToPagesWide = 1
                ws_excel.PageSetup.FitToPagesTall = 1
                ws_excel.PageSetup.Orientation = 1 # xlPortrait
                
                # Export to PDF (0 = xlTypePDF)
                excel_wb.ExportAsFixedFormat(0, abs_pdf_path)
                excel_wb.Close(False)
            except Exception as e:
                print(f"Warning: Failed to create PDF for {sheet_name}: {e}")
                
    if excel_app:
        try:
            excel_app.Quit()
        except:
            pass
        
    print(f"✅ Generated individual Salary Slips inside folder {output_dir}/")
    if excel_app:
        print(f"✅ Generated PDF Salary Slips inside folder {pdf_dir}/")

if __name__ == "__main__":
    def get_input(prompt_text, default_val):
        user_input = input(f"{prompt_text} [{default_val}]: ").strip()
        return user_input if user_input else default_val

    current_month_str = get_input("Enter month (e.g., Jan, Feb...)", datetime.today().strftime("%b"))
    current_year = get_input("Enter year (e.g., 2026)", str(datetime.today().year))

    master_attendance_file = "Attendance_Master.xlsx"
    
    if not os.path.exists(master_attendance_file):
        print(f"File {master_attendance_file} not found. Generating dummy dataset...")
        generate_dummy_data(master_attendance_file)
        
    print(f"Processing Master Data from {master_attendance_file}...")
    try:
        calculated_df = process_payroll(master_attendance_file, current_month_str, current_year)
        generate_salary_slips(calculated_df, current_month_str, current_year)
        print("Success! All operations completed.")
    except Exception as e:
        print("An error occurred:", e)
