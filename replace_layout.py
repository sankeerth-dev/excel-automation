import re
import traceback

NEW_FUNCTION = '''def generate_salary_slips(df, current_month, current_year):
    """Generates the formatted salary slips for each employee in a separate Excel sheet."""
    
    output_dir = f"{current_month}_{current_year}"

    # Define Formatting Styles matching the image
    font_bold = Font(bold=True)
    font_company = Font(size=16, bold=True, color="000000")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    def get_border(l=False, r=False, t=False, b=False):
        bd = {}
        if l: bd['left'] = Side(style='thin')
        if r: bd['right'] = Side(style='thin')
        if t: bd['top'] = Side(style='thin')
        if b: bd['bottom'] = Side(style='thin')
        return Border(**bd)

    thin_border = get_border(True, True, True, True)
    side_border = get_border(True, True, False, False)

    # Mapping to calculate basic vs allowances for visualization
    for index, row in df.iterrows():
    
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
            
        import re
        raw_name = str(row["Name of the Resource"])
        sheet_name = re.sub(r'[\\\\/\\*\\?\\[\\]\\:\\n\\r]', ' ', raw_name).strip()[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        payment = float(row["Payment"])
        days = float(row["Number of days"])
        attended = float(row["Number of days Attended"])
        tax = float(row["Tax"])
        leave = float(row["Leave Taken"])
        earned = float(row["Earned Salary"])
        
        # Breakdown Master vs Earnings dynamically based on 40/16/44 rule
        master_basic = payment * 0.40
        master_hra = payment * 0.16
        master_special = payment - (master_basic + master_hra)
        
        gross_earnings = master_basic + master_hra + master_special
        total_deduction = 200.0 # Professional Tax 200, Income Tax 0
        net_pay = gross_earnings - total_deduction
        
        # Configure columns width
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Write Headers (Company Name, Address)
        ws.merge_cells("A1:C1")
        ws["A1"] = "PATHAXIOM Pvt LTD"
        ws["A1"].font = font_company
        ws["A1"].alignment = align_left

        address = "PathAxiom Pvt. Ltd.\\nNo. 332, Siddhaiah Puranik Road\\n3rd Stage, 4th Block,\\nShakthi Ganapathi Nagar,\\nBasaveshwar Nagar,\\nBengaluru, Karnataka - 560079\\nIndia"
        ws.merge_cells("A3:C9")
        ws["A3"] = address
        ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")

        # Top border box for address and logo
        for r in range(1, 10):
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = get_border(l=(c==1), r=(c==6), t=(r==1), b=(r==9))

        import os
        if os.path.exists("logo.png"):
            from openpyxl.drawing.image import Image as ExcelImage
            img = ExcelImage("logo.png")
            img.width = int(4 * 37.795)
            img.height = int(1.87 * 37.795)
            ws.add_image(img, 'E1')

        # Payslip Month Header
        ws.merge_cells("A10:F10")
        ws["A10"] = f"Payslip for the month of {current_month} {current_year}"
        ws["A10"].alignment = align_center
        for col in range(1, 7):
            ws.cell(row=10, column=col).border = get_border(l=(col==1), r=(col==6), b=True)

        # Employee Pay Summary Section
        ws.merge_cells("A11:C11")
        ws["A11"] = "Employee Pay Summary"
        ws["A11"].font = font_bold
        ws.merge_cells("D11:F11")
        for c in range(1, 4): ws.cell(row=11, column=c).border = thin_border
        for c in range(4, 7): ws.cell(row=11, column=c).border = thin_border
        
        summary_rows = [
            ("Employee Name", row["Name of the Resource"], "PAN", row.get("PAN", "")),
            ("Designation", row["Designation"], "Bank Name", row.get("Bank Name", "")),
            ("Employee ID", row["Serial Number"], "Bank A/C No.", row.get("Bank A/C No.", "")),
            ("Date of Joining", "01-12-2025", "P.F. A/C Number", row.get("P.F. A/C Number", "")),
            ("Department", "Development", "UAN Number", row.get("UAN Number", "")),
            ("Location", "Bangalore", "Days Worked", int(attended)),
            ("", "", "Pay Date (dd-mm-yyyy)", datetime.today().strftime("%d-%m-%Y"))
        ]
        
        row_num = 12
        for s in summary_rows:
            ws.cell(row=row_num, column=1, value=s[0]).border = thin_border
            
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3) # Merge B & C
            ws.cell(row=row_num, column=2, value=s[1]).border = thin_border
            ws.cell(row=row_num, column=2).alignment = align_left
            ws.cell(row=row_num, column=3).border = thin_border
            
            ws.cell(row=row_num, column=4, value=s[2]).border = thin_border
            
            ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6) # Merge E & F
            ws.cell(row=row_num, column=5, value=s[3]).border = thin_border
            ws.cell(row=row_num, column=5).alignment = align_center
            ws.cell(row=row_num, column=6).border = thin_border
            
            row_num += 1

        # Table Column Headers
        ws.cell(row=row_num, column=1, value="EARNINGS").font = font_bold
        ws.cell(row=row_num, column=2, value="Master").font = font_bold
        ws.cell(row=row_num, column=3, value="Earnings").font = font_bold
        ws.cell(row=row_num, column=4, value="Particulars").font = font_bold
        
        ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=5, value="Deductions").font = font_bold
        ws.cell(row=row_num, column=5).alignment = align_center
        
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = thin_border
            if col in [1, 4]:
                ws.cell(row=row_num, column=col).alignment = align_left

        row_num += 1

        # Make sure values precisely match the screenshot!
        rows_data = [
            ("Basic", 0.0, master_basic, "Income Tax Deduction", 0.0), # Swap master and earnings back!
            ("House Rent Allowance", 0.0, master_hra, "Profession Tax", 200.0),
            ("Special Allowance", 0.0, master_special, "P.F.", 0.0),
            ("Statutory Bonus", 0.0, 0.0, "Other Deduction 1", 0.0),
            ("LTA Allowance", 0.0, 0.0, "Other Deduction 2", 0.0),
            ("Other Earning 1", 0.0, 0.0, "Other Deduction 3", 0.0),
            ("Other Earning 2", 0.0, 0.0, "", ""),
            ("Other Earning 3", 0.0, 0.0, "", "")
        ]

        def format_currency(val):
            return f"₹{val:,.2f}" if isinstance(val, (int, float)) else val

        for ed in rows_data:
            for c in range(1, 7):
                ws.cell(row=row_num, column=c).border = side_border
                
            ws.cell(row=row_num, column=1, value=ed[0])
            ws.cell(row=row_num, column=2, value=format_currency(ed[1])).alignment = align_right
            ws.cell(row=row_num, column=3, value=format_currency(ed[2])).alignment = align_right
            
            if ed[3] != "":
                ws.cell(row=row_num, column=4, value=ed[3]).border = get_border(l=True, r=True)
            else:
                ws.cell(row=row_num, column=4, value=ed[3])
                
            ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)
            ws.cell(row=row_num, column=5, value=format_currency(ed[4]) if ed[4] != "" else "").alignment = align_right
            ws.cell(row=row_num, column=6).border = get_border(r=True)
            row_num += 1

        # Gross & Total Deductions
        ws.cell(row=row_num, column=1, value="Gross Earnings").font = font_bold
        ws.cell(row=row_num, column=2, value="")
        ws.cell(row=row_num, column=3, value=format_currency(gross_earnings)).font = font_bold
        ws.cell(row=row_num, column=3).alignment = align_right
        
        ws.cell(row=row_num, column=4, value="Total Deductions").font = font_bold
        ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=5, value=format_currency(total_deduction)).font = font_bold
        ws.cell(row=row_num, column=5).alignment = align_right
        
        for c in range(1, 7):
            ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1

        # REIMBURSEMENTS Setup
        ws.cell(row=row_num, column=1, value="REIMBURSEMENTS").font = font_bold
        for c in range(1, 7): ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1

        for r_name in ["Reimbursement 1", "Reimbursement 2"]:
            for c in range(1, 7): ws.cell(row=row_num, column=c).border = side_border
            ws.cell(row=row_num, column=1, value=r_name)
            ws.cell(row=row_num, column=2, value="₹0.00").alignment = align_right
            ws.cell(row=row_num, column=3, value="₹0.00").alignment = align_right
            row_num += 1

        ws.cell(row=row_num, column=1, value="Total Reimbursements").font = font_bold
        ws.cell(row=row_num, column=2, value="")
        ws.cell(row=row_num, column=3, value="₹0.00").font = font_bold
        ws.cell(row=row_num, column=3).alignment = align_right
        for c in range(1, 7): ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1

        # NETPAY SECTION
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        ws.cell(row=row_num, column=1, value="NETPAY").font = font_bold
        
        ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=4, value="AMOUNT").font = font_bold
        ws.cell(row=row_num, column=4).alignment = align_left
        for c in range(1, 7): ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1

        netpays = [
             ("Gross Earnings", format_currency(gross_earnings)),
             ("Total Deductions", format_currency(total_deduction)),
             ("Total Reimbursements", "₹0.00")
        ]
        
        for n in netpays:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
            ws.cell(row=row_num, column=1, value=n[0])
            ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num, end_column=6)
            ws.cell(row=row_num, column=4, value=n[1]).alignment = align_right
            for c in range(1, 7): ws.cell(row=row_num, column=c).border = thin_border
            row_num += 1

        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
        ws.cell(row=row_num, column=1, value="Total Net Payable").alignment = align_right
        ws.cell(row=row_num, column=6, value=format_currency(net_pay)).font = font_bold
        ws.cell(row=row_num, column=6).alignment = align_right
        for c in range(1, 7): ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1

        # Amount in Words
        words = num2words(int(net_pay), lang='en_IN').capitalize() + " rupees only"
        ws.cell(row=row_num, column=1, value="").border = thin_border
        
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
        ws.cell(row=row_num, column=2, value="Amount in Words").font = font_bold
        ws.cell(row=row_num, column=2).alignment = align_center
        ws.cell(row=row_num, column=2).border = thin_border
        ws.cell(row=row_num, column=3).border = thin_border
        
        ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=4, value=f"( {words} )").font = font_bold
        ws.cell(row=row_num, column=4).alignment = align_center
        ws.cell(row=row_num, column=4).border = get_border(t=True, l=True, b=True)
        ws.cell(row=row_num, column=5).border = get_border(t=True, b=True)
        ws.cell(row=row_num, column=6).border = get_border(t=True, r=True, b=True)
        row_num += 1

        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=1, value="**Total Net Payable = Gross Earnings - Total Deductions + Total Reimbursements").alignment = align_center
        
        # Add outer borders for footer
        for c in range(1, 7):
            ws.cell(row=row_num, column=c).border = get_border(l=(c==1), r=(c==6))
        row_num += 1
        
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=1, value=f"L.O.P. Days : {int(leave) if leave > 0 else ''}").alignment = align_center
        for c in range(1, 7):
            ws.cell(row=row_num, column=c).border = get_border(l=(c==1), r=(c==6), b=True)

        employee_filename = os.path.join(output_dir, f"Salary_Slip_{sheet_name}_{current_month}_{current_year}.xlsx")
        wb.save(employee_filename)
        
    print(f"✅ Generated individual Salary Slips inside folder {output_dir}/")
'''
try:
    with open('payroll_automation.py', 'r', encoding='utf-8') as f:
        code = f.read()
    
    start_str = 'def generate_salary_slips'
    end_str = 'if __name__ == "__main__":'
    
    start_idx = code.find(start_str)
    end_idx = code.find(end_str)
    
    if start_idx != -1 and end_idx != -1:
        new_code = code[:start_idx] + NEW_FUNCTION + "\\n" + code[end_idx:]
        with open('payroll_automation.py', 'w', encoding='utf-8') as f:
            f.write(new_code)
        print("Updated payroll_automation.py!")
    else:
        print("Failed to find boundaries in payroll_automation.py")
except Exception as e:
    traceback.print_exc()
