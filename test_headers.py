import openpyxl
wb = openpyxl.load_workbook('Attendance_Master.xlsx')
ws = wb.active
columns = [(col, ws.cell(row=1, column=col).value) for col in range(1, ws.max_column + 1)]
for col in columns:
    print(col)
